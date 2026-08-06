"""Datová vrstva oddílů: import Excelu, trvalé uložení, export a sdílená jména."""

from __future__ import annotations

import json
import copy
import math
import os
import re
import tempfile
import unicodedata
import uuid
from datetime import date, datetime, timezone

from app_paths import get_user_data_dir


SCHEMA_VERSION = 1


class GroupsDataError(RuntimeError):
    """Srozumitelná chyba importu nebo exportu dat oddílů."""


def _uid() -> str:
    return uuid.uuid4().hex[:12]


def _plain_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Ano" if value else "Ne"
    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).strip().split())


def normalized_label(value) -> str:
    text = unicodedata.normalize("NFKD", _plain_text(value)).casefold()
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text)


def empty_groups_data() -> dict:
    return {
        "schema_version": SCHEMA_VERSION,
        "source": {},
        "groups": [],
    }


def groups_data_path() -> str:
    return os.path.join(get_user_data_dir("groups"), "oddily.json")


def load_groups_data(path: str | None = None) -> dict:
    target = path or groups_data_path()
    if not os.path.exists(target):
        return empty_groups_data()
    try:
        with open(target, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (OSError, ValueError, TypeError):
        return empty_groups_data()
    if not isinstance(data, dict) or not isinstance(data.get("groups"), list):
        return empty_groups_data()
    data.setdefault("schema_version", SCHEMA_VERSION)
    data.setdefault("source", {})
    return data


def save_groups_data(data: dict, path: str | None = None) -> str:
    target = os.path.abspath(path or groups_data_path())
    os.makedirs(os.path.dirname(target), exist_ok=True)
    payload = dict(data or {})
    payload["schema_version"] = SCHEMA_VERSION
    payload.setdefault("source", {})
    payload.setdefault("groups", [])
    file_descriptor, temporary = tempfile.mkstemp(
        prefix="oddily_", suffix=".json.tmp", dir=os.path.dirname(target)
    )
    try:
        with os.fdopen(file_descriptor, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, target)
    except Exception:
        try:
            os.unlink(temporary)
        except OSError:
            pass
        raise
    return target


def _xlsx_rows(path: str):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise GroupsDataError(
            "Pro import XLSX chybí knihovna openpyxl. Nainstalujte závislosti aplikace."
        ) from error
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise GroupsDataError(f"Soubor XLSX se nepodařilo otevřít: {error}") from error
    try:
        for sheet in workbook.worksheets:
            yield sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _xls_rows(path: str):
    try:
        import xlrd
    except ImportError as error:
        try:
            # Záložní kopie je součástí aplikace, aby XLS fungovalo také po
            # aktualizaci starších instalací a při spuštění bez pip instalace.
            from vendor import xlrd
        except ImportError:
            raise GroupsDataError(
                "Starší soubor XLS nelze otevřít. V instalaci chybí přibalená knihovna xlrd."
            ) from error
    try:
        workbook = xlrd.open_workbook(path, on_demand=True)
    except Exception as error:
        raise GroupsDataError(f"Soubor XLS se nepodařilo otevřít: {error}") from error
    try:
        for sheet in workbook.sheets():
            yield sheet.name, [sheet.row_values(row) for row in range(sheet.nrows)]
    finally:
        workbook.release_resources()


def _workbook_rows(path: str):
    extension = os.path.splitext(path)[1].casefold()
    if extension == ".xls":
        yield from _xls_rows(path)
        return
    if extension in (".xlsx", ".xlsm"):
        yield from _xlsx_rows(path)
        return
    raise GroupsDataError("Vyberte soubor Excel ve formátu .xls nebo .xlsx.")


def _unique_headers(row: list) -> list[tuple[int, str]]:
    result = []
    used = set()
    for column, value in enumerate(row):
        base = _plain_text(value)
        if not base:
            continue
        label = base
        suffix = 2
        while normalized_label(label) in used:
            label = f"{base} {suffix}"
            suffix += 1
        used.add(normalized_label(label))
        result.append((column, label))
    return result


def _row_value(row: list, column: int):
    return row[column] if 0 <= column < len(row) else None


def _person_from_row(row: list, headers: list[tuple[int, str]], role: str) -> dict | None:
    fields = {
        label: _plain_text(_row_value(row, column))
        for column, label in headers
    }
    if not any(fields.values()):
        return None
    name_fields = {
        normalized_label(label): value
        for label, value in fields.items()
    }
    if not (name_fields.get("jmeno") or name_fields.get("prijmeni")):
        return None
    return {"id": _uid(), "role": role, "fields": fields}


def _leader_from_row(row: list, first_name_column: int, surname_column: int) -> dict | None:
    first_name = _plain_text(_row_value(row, first_name_column))
    surname = _plain_text(_row_value(row, surname_column))
    if not first_name and not surname:
        return None
    return {
        "id": _uid(),
        "role": "leader",
        "fields": {"Jméno": first_name, "Příjmení": surname},
    }


def _find_header_row(rows: list[list]) -> int | None:
    for index, row in enumerate(rows):
        keys = {normalized_label(value) for value in row if _plain_text(value)}
        if "jmeno" in keys and "prijmeni" in keys:
            return index
    return None


def _find_leader_row(rows: list[list], start: int) -> int | None:
    for index in range(start, len(rows)):
        if any(normalized_label(value).startswith("vedouci") for value in rows[index]):
            return index
    return None


def _group_name(sheet_name: str, rows: list[list], header_index: int) -> str:
    for row in rows[:header_index]:
        for value in row:
            text = _plain_text(value)
            if text and "oddil" in normalized_label(text):
                return text
    return _plain_text(sheet_name) or "Oddíl"


def parse_groups_workbook(path: str) -> dict:
    source = os.path.abspath(path)
    if not os.path.isfile(source):
        raise GroupsDataError("Vybraný soubor neexistuje.")

    groups = []
    skipped_sheets = []
    for sheet_name, rows in _workbook_rows(source):
        header_index = _find_header_row(rows)
        if header_index is None:
            skipped_sheets.append(sheet_name)
            continue
        headers = _unique_headers(rows[header_index])
        leader_index = _find_leader_row(rows, header_index + 1)
        children_end = leader_index if leader_index is not None else len(rows)
        children = []
        for row in rows[header_index + 1:children_end]:
            person = _person_from_row(row, headers, "child")
            if person is not None:
                children.append(person)

        leaders = []
        if leader_index is not None:
            marker_row = rows[leader_index]
            marker_column = next(
                (
                    column
                    for column, value in enumerate(marker_row)
                    if normalized_label(value).startswith("vedouci")
                ),
                0,
            )
            first_name_column = marker_column + 1
            surname_column = marker_column + 2
            for row in rows[leader_index:]:
                leader = _leader_from_row(row, first_name_column, surname_column)
                if leader is not None:
                    leaders.append(leader)

        if children or leaders:
            groups.append(
                {
                    "id": _uid(),
                    "name": _group_name(sheet_name, rows, header_index),
                    "children": children,
                    "leaders": leaders,
                }
            )

    if not groups:
        detail = f" Přeskočené listy: {', '.join(skipped_sheets)}." if skipped_sheets else ""
        raise GroupsDataError(
            "V souboru nebyl nalezen žádný list s hlavičkami Jméno a Příjmení." + detail
        )

    return {
        "schema_version": SCHEMA_VERSION,
        "source": {
            "file_name": os.path.basename(source),
            "imported_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "skipped_sheets": skipped_sheets,
        },
        "groups": groups,
    }


def _person_name_key(person: dict) -> str:
    return "|".join(
        (
            normalized_label(person_field(person, "Jméno")),
            normalized_label(person_field(person, "Příjmení")),
        )
    ).strip("|")


def _person_number_key(person: dict) -> str:
    return normalized_label(person_field(person, "Počet"))


def _merged_person(existing: dict, imported: dict) -> dict:
    """Aktualizuje importovaná pole a zachová ručně přidané údaje (např. telefon)."""
    result = {
        "id": existing.get("id") or imported.get("id") or _uid(),
        "role": imported.get("role") or existing.get("role") or "child",
        "fields": dict(existing.get("fields") or {}),
    }
    normalized_existing = {
        normalized_label(label): label for label in result["fields"]
    }
    for label, value in (imported.get("fields") or {}).items():
        key = normalized_label(label)
        target_label = normalized_existing.get(key, str(label))
        result["fields"][target_label] = _plain_text(value)
        normalized_existing[key] = target_label
    return result


def _merge_people(existing_people: list[dict], imported_people: list[dict]) -> tuple[list[dict], dict]:
    result = []
    used_indexes = set()
    stats = {"added": 0, "updated": 0, "unchanged": 0, "retained": 0}

    def match_index(imported_person: dict) -> int | None:
        name_key = _person_name_key(imported_person)
        number_key = _person_number_key(imported_person)
        if name_key:
            for index, current in enumerate(existing_people):
                if index not in used_indexes and _person_name_key(current) == name_key:
                    return index
        if number_key:
            for index, current in enumerate(existing_people):
                if index not in used_indexes and _person_number_key(current) == number_key:
                    return index
        return None

    for imported_person in imported_people:
        index = match_index(imported_person)
        if index is None:
            result.append(copy.deepcopy(imported_person))
            stats["added"] += 1
            continue
        used_indexes.add(index)
        current = existing_people[index]
        merged = _merged_person(current, imported_person)
        if merged == current:
            stats["unchanged"] += 1
        else:
            stats["updated"] += 1
        result.append(merged)

    # Ručně přidané osoby se při synchronizaci neztratí. Odstranit je lze
    # z jejich osobní karty, což je bezpečnější než tiché smazání importem.
    for index, current in enumerate(existing_people):
        if index not in used_indexes:
            result.append(copy.deepcopy(current))
            stats["retained"] += 1
    return result, stats


def merge_groups_data(existing: dict, imported: dict) -> tuple[dict, dict]:
    """Sloučí opakovaný import bez duplikování oddílů a osob.

    Oddíly se párují podle názvu. Osoby nejdříve podle jména a příjmení,
    při přejmenování také podle sloupce ``Počet``. Importovaná pole se
    aktualizují, zatímco vlastní ručně přidaná pole zůstanou zachována.
    """
    current_groups = list((existing or {}).get("groups") or [])
    incoming_groups = list((imported or {}).get("groups") or [])
    by_name = {}
    for index, group in enumerate(current_groups):
        by_name.setdefault(normalized_label(group.get("name")), []).append(index)

    used_groups = set()
    merged_groups = []
    stats = {
        "groups_added": 0,
        "groups_updated": 0,
        "groups_retained": 0,
        "people_added": 0,
        "people_updated": 0,
        "people_unchanged": 0,
        "people_retained": 0,
    }

    for incoming in incoming_groups:
        key = normalized_label(incoming.get("name"))
        candidates = by_name.get(key, [])
        current_index = next((idx for idx in candidates if idx not in used_groups), None)
        if current_index is None:
            merged_groups.append(copy.deepcopy(incoming))
            stats["groups_added"] += 1
            stats["people_added"] += len(incoming.get("children", [])) + len(incoming.get("leaders", []))
            continue

        used_groups.add(current_index)
        current = current_groups[current_index]
        merged_group = {
            "id": current.get("id") or incoming.get("id") or _uid(),
            "name": _plain_text(incoming.get("name")) or _plain_text(current.get("name")) or "Oddíl",
        }
        changed = merged_group["name"] != current.get("name")
        for collection in ("children", "leaders"):
            people, people_stats = _merge_people(
                list(current.get(collection) or []),
                list(incoming.get(collection) or []),
            )
            merged_group[collection] = people
            changed = changed or people != list(current.get(collection) or [])
            stats["people_added"] += people_stats["added"]
            stats["people_updated"] += people_stats["updated"]
            stats["people_unchanged"] += people_stats["unchanged"]
            stats["people_retained"] += people_stats["retained"]
        merged_groups.append(merged_group)
        if changed:
            stats["groups_updated"] += 1

    for index, group in enumerate(current_groups):
        if index not in used_groups:
            merged_groups.append(copy.deepcopy(group))
            stats["groups_retained"] += 1
            stats["people_retained"] += len(group.get("children", [])) + len(group.get("leaders", []))

    payload = {
        "schema_version": SCHEMA_VERSION,
        "source": dict((imported or {}).get("source") or {}),
        "groups": merged_groups,
    }
    payload["source"]["merge_stats"] = dict(stats)
    return payload, stats


def person_field(person: dict, label: str) -> str:
    target = normalized_label(label)
    for current_label, value in (person.get("fields") or {}).items():
        if normalized_label(current_label) == target:
            return _plain_text(value)
    return ""


def person_display_name(person: dict) -> str:
    first_name = person_field(person, "Jméno")
    surname = person_field(person, "Příjmení")
    name = " ".join(part for part in (first_name, surname) if part).strip()
    if name:
        return name
    values = [_plain_text(value) for value in (person.get("fields") or {}).values()]
    return next((value for value in values if value), "Neznámá osoba")


def average_child_age(group: dict) -> float | None:
    ages = []
    for child in group.get("children", []):
        raw = person_field(child, "Věk").replace(",", ".")
        try:
            age = float(raw)
        except (TypeError, ValueError):
            continue
        if math.isfinite(age) and age >= 0:
            ages.append(age)
    return sum(ages) / len(ages) if ages else None


def roster_entries(data: dict | None = None) -> list[dict]:
    payload = data if data is not None else load_groups_data()
    entries = []
    for group in payload.get("groups", []):
        for role_key, role_label in (("leaders", "Vedoucí"), ("children", "Dítě")):
            for person in group.get(role_key, []):
                name = person_display_name(person)
                entries.append(
                    {
                        "name": name,
                        "group_name": _plain_text(group.get("name")),
                        "group_id": group.get("id"),
                        "person_id": person.get("id"),
                        "role": role_label,
                        "fields": dict(person.get("fields") or {}),
                    }
                )
    return entries


def _field_union(people: list[dict]) -> list[str]:
    result = []
    known = set()
    for person in people:
        for label in (person.get("fields") or {}).keys():
            key = normalized_label(label)
            if key and key not in known:
                known.add(key)
                result.append(str(label))
    return result


def _safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", _plain_text(name))[:31] or "Oddíl"
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        marker = f" {suffix}"
        candidate = base[:31 - len(marker)] + marker
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def export_groups_workbook(groups: list[dict], path: str) -> str:
    if not groups:
        raise GroupsDataError("Není co exportovat. Nejprve importujte alespoň jeden oddíl.")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise GroupsDataError(
            "Pro export do Excelu chybí knihovna openpyxl. Nainstalujte závislosti aplikace."
        ) from error

    target = os.path.abspath(path)
    if not target.casefold().endswith(".xlsx"):
        target += ".xlsx"
    os.makedirs(os.path.dirname(target), exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names = set()
    navy = "0B2230"
    teal = "0E4C57"
    gold = "C99B4E"
    pale = "FFF7E4"
    thin_gold = Side(style="thin", color=gold)

    for group in groups:
        sheet = workbook.create_sheet(_safe_sheet_name(group.get("name", "Oddíl"), used_names))
        children = list(group.get("children", []))
        leaders = list(group.get("leaders", []))
        child_headers = _field_union(children) or ["Jméno", "Příjmení", "Věk"]
        leader_headers = _field_union(leaders) or ["Jméno", "Příjmení"]
        column_count = max(2, len(child_headers), len(leader_headers))

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        title = sheet.cell(1, 1, _plain_text(group.get("name")) or "Oddíl")
        title.fill = PatternFill("solid", fgColor=navy)
        title.font = Font(name="Georgia", size=20, bold=True, color="F4DEA4")
        title.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 34

        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=column_count)
        child_title = sheet.cell(3, 1, f"DĚTI • {len(children)}")
        child_title.fill = PatternFill("solid", fgColor=teal)
        child_title.font = Font(name="Georgia", bold=True, color="FFFFFF")

        for column, label in enumerate(child_headers, 1):
            cell = sheet.cell(4, column, label)
            cell.fill = PatternFill("solid", fgColor=gold)
            cell.font = Font(bold=True, color=navy)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin_gold)
        for row_index, person in enumerate(children, 5):
            fields = person.get("fields") or {}
            lookup = {normalized_label(label): value for label, value in fields.items()}
            for column, label in enumerate(child_headers, 1):
                cell = sheet.cell(row_index, column, _plain_text(lookup.get(normalized_label(label), "")))
                cell.fill = PatternFill("solid", fgColor=pale if row_index % 2 else "F3EAD4")
                cell.border = Border(bottom=Side(style="hair", color="D9C28B"))

        leader_title_row = 6 + len(children)
        sheet.merge_cells(
            start_row=leader_title_row,
            start_column=1,
            end_row=leader_title_row,
            end_column=column_count,
        )
        leader_title = sheet.cell(leader_title_row, 1, f"VEDOUCÍ • {len(leaders)}")
        leader_title.fill = PatternFill("solid", fgColor=teal)
        leader_title.font = Font(name="Georgia", bold=True, color="FFFFFF")
        for column, label in enumerate(leader_headers, 1):
            cell = sheet.cell(leader_title_row + 1, column, label)
            cell.fill = PatternFill("solid", fgColor=gold)
            cell.font = Font(bold=True, color=navy)
            cell.alignment = Alignment(horizontal="center")
        for offset, person in enumerate(leaders, 2):
            fields = person.get("fields") or {}
            lookup = {normalized_label(label): value for label, value in fields.items()}
            for column, label in enumerate(leader_headers, 1):
                cell = sheet.cell(
                    leader_title_row + offset,
                    column,
                    _plain_text(lookup.get(normalized_label(label), "")),
                )
                cell.fill = PatternFill("solid", fgColor=pale if offset % 2 else "F3EAD4")

        sheet.freeze_panes = "A5"
        sheet.sheet_view.showGridLines = False
        for column in range(1, column_count + 1):
            values = [
                _plain_text(sheet.cell(row, column).value)
                for row in range(1, sheet.max_row + 1)
            ]
            width = min(42, max(12, max((len(value) for value in values), default=0) + 3))
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.auto_filter.ref = f"A4:{sheet.cell(4 + len(children), len(child_headers)).coordinate}"

    try:
        workbook.save(target)
    except Exception as error:
        raise GroupsDataError(f"Excel se nepodařilo uložit: {error}") from error
    return target
