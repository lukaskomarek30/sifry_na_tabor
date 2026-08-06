import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from groups_data import (
    average_child_age,
    export_groups_workbook,
    merge_groups_data,
    parse_groups_workbook,
    person_display_name,
    roster_entries,
)


class GroupsDataTests(unittest.TestCase):
    def _build_workbook(self, path):
        workbook = Workbook()
        first = workbook.active
        first.title = "Modří"
        second = workbook.create_sheet("Zlatí")
        for index, (sheet, count) in enumerate(((first, 2), (second, 4)), 1):
            sheet.append([f"{index}. oddíl"])
            sheet.append([])
            sheet.append(["Počet", "Příjmení", "Jméno", "Věk", "Oddíl", "Ubytování", "Poznámka"])
            for child_index in range(count):
                sheet.append(
                    [
                        child_index + 1,
                        f"Příjmení {child_index}",
                        f"Dítě {index}",
                        8 + child_index,
                        index,
                        "Kajuta",
                        f"poznámka {child_index}",
                    ]
                )
            sheet.append([])
            sheet.append(["Vedoucí:", f"Vedoucí {index}", "Kapitán"])
            sheet.append([None, f"Pomocník {index}", "Kormidelník"])
        workbook.save(path)

    def test_variable_group_sizes_and_extra_columns(self):
        with tempfile.TemporaryDirectory() as folder:
            source = os.path.join(folder, "oddily.xlsx")
            self._build_workbook(source)
            data = parse_groups_workbook(source)

            self.assertEqual(2, len(data["groups"]))
            self.assertEqual([2, 4], [len(group["children"]) for group in data["groups"]])
            self.assertEqual([2, 2], [len(group["leaders"]) for group in data["groups"]])
            self.assertEqual("Dítě 1 Příjmení 0", person_display_name(data["groups"][0]["children"][0]))
            self.assertEqual("poznámka 0", data["groups"][0]["children"][0]["fields"]["Poznámka"])
            self.assertEqual(8.5, average_child_age(data["groups"][0]))
            self.assertEqual(9.5, average_child_age(data["groups"][1]))

    def test_custom_leader_field_survives_export(self):
        with tempfile.TemporaryDirectory() as folder:
            source = os.path.join(folder, "oddily.xlsx")
            target = os.path.join(folder, "export.xlsx")
            self._build_workbook(source)
            data = parse_groups_workbook(source)
            data["groups"][0]["leaders"][0]["fields"]["Telefon"] = "+420 111 222 333"
            export_groups_workbook(data["groups"], target)

            workbook = load_workbook(target, read_only=True, data_only=True)
            try:
                values = [
                    tuple(value for value in row)
                    for row in workbook[workbook.sheetnames[0]].iter_rows(values_only=True)
                ]
            finally:
                workbook.close()
            self.assertTrue(any("Telefon" in row for row in values))
            self.assertTrue(any("+420 111 222 333" in row for row in values))

    def test_roster_entries_contain_children_and_leaders(self):
        with tempfile.TemporaryDirectory() as folder:
            source = os.path.join(folder, "oddily.xlsx")
            self._build_workbook(source)
            data = parse_groups_workbook(source)
            entries = roster_entries(data)
            self.assertEqual(10, len(entries))
            self.assertEqual({"Dítě", "Vedoucí"}, {entry["role"] for entry in entries})
            self.assertTrue(all(entry["group_name"] for entry in entries))

    def test_repeated_import_updates_without_duplicates_and_keeps_custom_fields(self):
        with tempfile.TemporaryDirectory() as folder:
            source = os.path.join(folder, "oddily.xlsx")
            self._build_workbook(source)
            original = parse_groups_workbook(source)
            first_child = original["groups"][0]["children"][0]
            original_id = first_child["id"]
            first_child["fields"]["Telefon rodiče"] = "+420 777 000 111"

            changed = parse_groups_workbook(source)
            changed["groups"][0]["children"][0]["fields"]["Věk"] = "12"
            changed["groups"][0]["children"].append(
                {
                    "id": "new-child",
                    "role": "child",
                    "fields": {"Počet": "99", "Jméno": "Nové", "Příjmení": "Dítě", "Věk": "9"},
                }
            )

            merged, stats = merge_groups_data(original, changed)
            children = merged["groups"][0]["children"]
            self.assertEqual(3, len(children))
            self.assertEqual(original_id, children[0]["id"])
            self.assertEqual("12", children[0]["fields"]["Věk"])
            self.assertEqual("+420 777 000 111", children[0]["fields"]["Telefon rodiče"])
            self.assertEqual(1, stats["people_added"])

            merged_again, _stats = merge_groups_data(merged, changed)
            self.assertEqual(3, len(merged_again["groups"][0]["children"]))


if __name__ == "__main__":
    unittest.main()
